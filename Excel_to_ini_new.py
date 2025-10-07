import openpyxl
import re
import configparser
import os
from datetime import datetime

# --- Helpers ---
def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    elif isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%d-%b-%y", "%d-%b-%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).strftime("%d-%m-%Y")
            except:
                pass
        return value
    return str(value) if value is not None else ""

def clean_number(value):
    """Convert cell to string number, remove alphabets but keep decimals as-is."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)  # keep original decimal formatting
    
    # Remove all alphabets and spaces, keep digits, dot, minus
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    
    if cleaned == "":
        return ""
    
    return cleaned

def collect(sheet, rule, key=None):
    """Extract values from sheet based on rule."""
    # --- Case 1: Single Cell (T4, W4, D14) ---
    if re.match(r"^[A-Z]+\d+$", rule):
        value = sheet[rule].value
        if key and "date" in key.lower():
            return format_date(value)
        return str(value or "")

    # --- Case 2: Column down (F38+, J38+, etc.) ---
    if rule.endswith("+"):
        col = re.findall(r"[A-Z]+", rule)[0]
        row = int(re.findall(r"\d+", rule)[0])
        values = []
        while True:
            cell_value = sheet[f"{col}{row}"].value
            if cell_value is None or str(cell_value).strip() == "":
                break
            clean_value = str(cell_value).replace("\n", " ").replace("\r", " ")
            clean_value = re.sub(r"\s+", " ", clean_value).strip()

            # --- special case for description ---
            if key and key.lower() == "description":
                words = clean_value.split(" ", 1)
                if len(words) > 1 and words[0].lower() == "s":
                    clean_value = words[1]

            # For numeric-only columns like sno, clean numbers
            if key and key.lower() in ("sno", "unit price", "bcd", "sws", "igst", "assess value", "total duty"):
                clean_value = clean_number(clean_value)

            values.append(clean_value)
            row += 1
        return ";".join(values)

    # --- Case 3: Step Rule (D24:step20) ---
    if ":step" in rule:
        col, rest = rule.split(":step")
        col_letter = re.findall(r"[A-Z]+", col)[0]
        row = int(re.findall(r"\d+", col)[0])
        step = int(rest)
        values = []
        while True:
            val = sheet[f"{col_letter}{row}"].value
            if val is None or str(val).strip() == "":
                break
            values.append(clean_number(val))
            row += step
        return ";".join(values)

    return ""

# --- Main Program ---
PAGE_TITLE_MAP = {
    "PART - I - BILL OF ENTRY SUMMARY": "PART_I",
    "PART - II - INVOICE &VALUATION DETAILS (Invoice 1 1 )": "PART_II",  # special handling below
    "PART - III - DUTIES": "PART_III",
    "PART - IV - ADDITIONAL DETAILS": "PART_IV",
    "PART - V - OTHER COMPLIANCES": "PART_V",
    "PART - VI - DECLARATION": "PART_VI",
}

def main():
    excel_file = os.path.join("Info", "Data_Extracted", "Extracted.xlsx")
    mapping_file = os.path.join("Info", "Config", "Mapping.ini")
    output_file = os.path.join("Info", "Config", "output.ini")

    # Load workbook
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheetnames = workbook.sheetnames

    # Load mapping
    config = configparser.ConfigParser()
    config.optionxform = str  # preserve case
    config.read(mapping_file)

    results = {}

    for idx, sheet_name in enumerate(sheetnames, start=1):
        sheet = workbook[sheet_name]
        title = str(sheet["A12"].value).strip() if sheet["A12"].value else ""
        title_upper = title.upper()
        title_norm = title_upper.replace(" ", "")  # normalize by removing spaces

        section = None
        mapping_section = None

        # --- Special handling for PART II (Invoices) ---
        if title_norm.startswith("PART-II-INVOICE&VALUATION"):
            mapping_section = "PART_II" if idx == 2 else "PART_II_Extra"
            section = "PART_II"   # merge all into PART_II

        elif title_norm.startswith("PART-III-DUTIES"):
            section = "PART_III"
            d32_val = sheet["D32"].value
            if d32_val is not None and str(d32_val).strip() != "":
                mapping_section = "PART_III_Extra_Col"
            else:
                mapping_section = "PART_III"

        elif title_norm.startswith("PART-I-BILLOFENTRYSUMMARY"):
            section = mapping_section = "PART_I"
        elif title_norm.startswith("PART-IV-ADDITIONALDETAILS"):
            section = mapping_section = "PART_IV"
        elif title_norm.startswith("PART-V-OTHERCOMPLIANCES"):
            section = mapping_section = "PART_V"
        elif title_norm.startswith("PART-VI-DECLARATION"):
            section = mapping_section = "PART_VI"

        # --- If no mapping found, skip ---
        if not mapping_section or mapping_section not in config.sections():
            print(f"[WARNING] Skipping {sheet_name}, no mapping for title: {title}")
            continue

        # Collect values
        if section not in results:
            results[section] = {}

        for key, rule in config[mapping_section].items():
            value = collect(sheet, rule, key)
            if value:
                if key not in results[section]:
                    results[section][key] = []
                results[section][key].append(value)

    # --- Write merged results to output.ini ---
    with open(output_file, "w", encoding="utf-8") as f:
        for section, kv in results.items():
            f.write(f"[{section}]\n")
            for key, values in kv.items():
                # Flatten all values into a single list, split on ';'
                flat_values = []
                for v in values:
                    flat_values.extend(v.split(";"))
                flat_values = [x for x in flat_values if x.strip()]  # remove empties
                merged = ";".join(flat_values)

                # Debug print for PART_II sno
                #if section == "PART_II" and key.lower() == "sno":
                    #print(f"DEBUG PART_II -> sno = {merged}")

                f.write(f"{key} = {merged}\n")
            f.write("\n")

    print(f"Extraction completed. Output saved to {output_file}")

if __name__ == "__main__":
    main()
