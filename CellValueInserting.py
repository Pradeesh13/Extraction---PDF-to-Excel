import configparser
import openpyxl
from decimal import Decimal, InvalidOperation
import sys
import os
from openpyxl.styles import Border, Side, Alignment, PatternFill

# --- File paths ---

ini_file = os.path.join("Info", "Config", "output.ini")
cell_map_file = os.path.join("Info", "Config", "Cell_Location.txt")
excel_file = os.path.join("Info", "Template", "Template.xlsx")

# Load output file path from path.ini
path_ini_file = os.path.join("Info", "Config", "path.ini")
path_config = configparser.ConfigParser()
path_config.read(path_ini_file)

output_file = path_config.get('output', 'path')

# --- Step 1: Read INI file ---

config = configparser.ConfigParser()
config.read(ini_file)

def normalize(key: str) -> str:
    return key.strip().lower().replace(" ", "")

def convert_value(val: str):
    """Convert to int or Decimal, preserving exact precision from INI."""
    val = val.strip()
    try:
        if "." not in val:  # pure integer
            return int(val)
        return Decimal(val)  # exact float with full precision
    except (ValueError, InvalidOperation):
        return val

ini_data = {}
for section in config.sections():
    for key, value in config.items(section):
        ini_data[normalize(key)] = value.strip()
if not ini_data:
    sys.exit(1)

# --- Step 2: Read Cell_Location.txt ---

cell_map = {}
with open(cell_map_file, "r") as f:
    for line in f:
        if "=" in line:
            key, cell = line.split("=", 1)
            key = normalize(key)
            cell = cell.strip().strip("()")
            row, col = map(int, cell.split(","))
            cell_map.setdefault(key, []).append((row, col))

# --- Step 3: Open Excel template ---

wb = openpyxl.load_workbook(excel_file)
ws = wb.active  # or ws = wb["YourSheetName"]

# --- Define border and alignment styles ---

thin = Side(border_style="thin", color="000000")
full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
center_align = Alignment(horizontal="center", vertical="center")
green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

# --- Step 4: Insert values ---

for key, locations in cell_map.items():
    if key in ini_data:
        value = ini_data[key]

        # Special case: Debit Duty -> sum of values
        if key == "debitduty":
            parts = [p.strip() for p in value.split(";") if p.strip()]
            try:
                total = sum(Decimal(p) for p in parts)
            except InvalidOperation:
                total = sum(float(p) for p in parts)
            for row, col in locations:
                cell = ws.cell(row=row, column=col, value=total)
                cell.border = full_border
                cell.alignment = center_align
                if parts and "." in parts[0]:
                    decimals = len(parts[0].split(".")[1])
                    cell.number_format = "0." + "0" * decimals
            continue

        for row, col in locations:
            if ";" in value:
                parts = [p.strip() for p in value.split(";") if p.strip()]
                for i, part in enumerate(parts):
                    cv = convert_value(part)
                    cell = ws.cell(row=row + i, column=col, value=cv)
                    cell.border = full_border
                    cell.alignment = center_align
                    if isinstance(cv, Decimal) and "." in part:
                        decimals = len(part.split(".")[1])
                        cell.number_format = "0." + "0" * decimals

                    # Always repeat invoice no next to unit price
                    if key == "unitprice" and "invoiceno" in ini_data:
                        invoice_val = ini_data["invoiceno"]
                        invoice_cell = ws.cell(row=row + i, column=col + 1, value=invoice_val)
                        invoice_cell.border = full_border
                        invoice_cell.alignment = center_align

            else:
                cv = convert_value(value)
                cell = ws.cell(row=row, column=col, value=cv)
                cell.border = full_border
                cell.alignment = center_align
                if isinstance(cv, Decimal) and "." in value:
                    decimals = len(value.split(".")[1])
                    cell.number_format = "0." + "0" * decimals

                # Single value case — repeat invoice no next to unit price
                if key == "unitprice" and "invoiceno" in ini_data:
                    invoice_val = ini_data["invoiceno"]
                    invoice_cell = ws.cell(row=row, column=col + 1, value=invoice_val)
                    invoice_cell.border = full_border
                    invoice_cell.alignment = center_align

# --- Step 4.1: Determine last sno and add summary section ---

sno_values = [p.strip() for p in ini_data.get("sno", "").split(";") if p.strip().isdigit()]
last_sno = int(sno_values[-1]) if sno_values else 0

# Summary section row is one row below the last sno row, plus 10 rows space
section_row = last_sno + 10

col = 15  # Column 'N' (1-based index, where A=1)

labels = ["Total Duty", "License", "As per BE Duty Amount", "Diff should be Nil"]

total_duty_parts = [p.strip() for p in ini_data.get("totalduty", "0").split(";") if p.strip()]
try:
    total_duty_sum = sum(Decimal(p) for p in total_duty_parts)
except InvalidOperation:
    total_duty_sum = sum(float(p) for p in total_duty_parts)

values = [total_duty_sum, 0, 0, 0]

for i, label in enumerate(labels):
    label_cell = ws.cell(row=section_row + i, column=col - 1, value=label)
    label_cell.border = full_border
    label_cell.alignment = center_align
    value_cell = ws.cell(row=section_row + i, column=col, value=values[i])
    value_cell.border = full_border
    value_cell.alignment = center_align
    if label == "Diff should be Nil":
        label_cell.fill = green_fill
        value_cell.fill = green_fill

# --- Step 5: Save ---

wb.save(output_file)

print("\033[1mSuccessful\033[0m")
